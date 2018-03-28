#!python3

# Author: Kevin Cruse
# Version: 1.1.0
# Description: Scrape bond data from Morningstar and dump into Excel
#
# Jargon:
#   Index bonds: Stable bonds used as a control to evaluate performance of non-index bonds
#   Non-index bonds: Bonds that aren't used as indexes (typically fluctuate more than index bonds)
#
# Shorthands:
#   TTM: Trailing twelve month
#   TTR: Trailing total return

import sys
import openpyxl
from openpyxl.styles import Font, colors
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait

# Aggregate storage of bond data found on Morningstar
class Bond:
    def __init__(self, ticker='', row=0):
        self.ticker = ticker  # Ticker name of bond
        self.row = row        # Row number of bond in spreadsheet
        self.index = []       # Column numbers of empty bond values on spreadsheet that need to be filled with index bond values
        self.exp = 0.0        # Expense ratio of bond
        self.yld = 0.0        # TTM yield of bond
        self.ytd = 0.0        # Year-to-date TTR of bond
        self.mtd = 0.0        # Month-to-date TTR of bond
        self.qtd = 0.0        # Quarter-to-date TTR of bond
        self.t1 = 0.0         # 1-year TTR of bond
        self.t3 = 0.0         # 3-year TTR of bond
        self.t5 = 0.0         # 5-year TTR of bond

# Get list of bonds or indexes from spreadsheet
def get_bonds(sheet=None, only_index=False):
    row = 0      # Iterator for traversing rows of spreadsheet
    ticker = ''  # Ticker name of bond    
    bonds = []   # List to store bonds found in spreadsheet
    
    for row in range(2, sheet.max_row + 1):
        # If row contains ticker, then add bond to list
        ticker = sheet.cell(row=row, column=3).value
        if ticker and len(ticker) <= 5:
            # If only index bonds are desired, check if bond is index before appending to list
            if only_index:
                if sheet.cell(row=row, column=14).value == None:
                    bonds.append(Bond(ticker, row))
            # Else, append bond to list
            else:
                if sheet.cell(row=row, column=14).value != None:
                    bonds.append(Bond(ticker, row))
    
    return bonds

# Get bond values from Morningstar
def get_values(browser=None, bonds=[]):
    bond = None      # Iterator for traversing list of bonds
    index = 1        # Index for compare list
    compare = [0, 1] # Store data from different loads to check for errors    
    failures = 0     # Number of failures loading Morningstar
    element = None   # HTML element
    i = 0            # General iterator
    yeet = True
    
    for bond in bonds:
        print('Fetching data for ' + bond.ticker + '...')

        # Temporary fix for stale element issue until new method is devised
        ripe = True
        while ripe:
            try:
                # Get TTM yield
                # Only compare results when page has been loaded multiple of 2 times
                index = 1
                compare = [0, 1]
                while not (index and compare[0] == compare[1]):
                    # Switch between first and second element of compare list
                    index = int(not index)
                    
                    # Load page
                    failures = 0
                    while failures != -1:
                        try:
                            element = []
                            browser.get('http://quotes.morningstar.com/fund/fundquote/f?t=' + bond.ticker + '&culture=en_us&platform=RET&test=QuoteiFrame')
                            element.append(WebDriverWait(browser, 20).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR, 'td[class="gr_table_colm2b"] > span'))))
                            element.append(WebDriverWait(browser, 20).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR, 'td[class="gr_table_colm21"] > span'))))                    
                            failures = -1
                        except Exception as e:
                            print('\nError loading page:\n' + e)
                        
                            # Exit program if page fails to load 5 times
                            failures += 1
                            if failures < 5:
                                print('Refreshing...\n')
                            else:
                                print('Unable to retreive data from Morningstar. Exiting program...')
                                sys.exit()
                    
                    # Store collected data as text and reset failure count
                    compare[index] = []
                    compare[index].append(element[0].text.rstrip('%'))
                    compare[index].append(element[1].text.rstrip('%'))

                ripe = False
            except Exception as e:
                    print('\nStale element\n')

        # Store and display TTM yield
        bond.exp = compare[0][0]
        print('Expense Ratio'.ljust(29) + ':' + bond.exp.rjust(7) + '%')
        bond.yld = compare[0][1]
        print('TTM Yield'.ljust(29) + ':' + bond.yld.rjust(7) + '%')

        ripe = True
        while ripe:
            try:
                # Get rest of bond data
                # Only compare results when page has been loaded multiple of 2 times
                index = 1
                compare = [0, 1]
                while not (index and compare[0] == compare[1]):
                    # Switch between first and second element of compare list
                    index = int(not index) 
                    
                    # Load page
                    failures = 0
                    while failures != -1:
                        try:
                            browser.get('http://performance.morningstar.com/fund/performance-return.action?t=' + bond.ticker + '&region=usa&culture=en_US')
                            element = WebDriverWait(browser, 20).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR, 'a[tabname="#tabquarter"]')))
                            element.click()
                            element = WebDriverWait(browser, 20).until(expected_conditions.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[id="tab-quar-end-content"] td[class="row_data"]')))
                            failures = -1
                        except Exception as e:
                            print('\nError loading page:\n')
                            print(e)

                            # Exit program if page fails to load 5 times
                            failures += 1
                            if failures < 5:
                                print('Refreshing...\n')
                            else:
                                print('Unable to retreive data from Morningstar. Exiting program...')
                                sys.exit()
                    
                    # Store collected data as text and reset failure count            
                    compare[index] = []
                    for i in range(len(element)):
                        compare[index].append(element[i].text)

                ripe = False
            except Exception as e:
                print('\nStale element\n')

        # Store and display rest of bond data
        bond.ytd = compare[0][3]
        print('YTD Trailing Total Return'.ljust(29) + ':' + bond.ytd.rjust(7) + '%')
        bond.mtd = compare[0][0]
        print('MTD Trailing Total Return'.ljust(29) + ':' + bond.mtd.rjust(7) + '%')
        bond.qtd = compare[0][1]
        print('QTD Trailing Total Return'.ljust(29) + ':' + bond.qtd.rjust(7) + '%')
        bond.t1 = compare[0][4]
        print('1-Year Trailing Total Return'.ljust(29) + ':' + bond.t1.rjust(7) + '%')
        bond.t3 = compare[0][5]
        print('3-Year Trailing Total Return'.ljust(29) + ':' + bond.t3.rjust(7) + '%')
        bond.t5 = compare[0][6]
        print('5-Year Trailing Total Return'.ljust(29) + ':' + bond.t5.rjust(7) + '%\n')

# Convert bond values from strings to floats
def to_floats(bonds=[]):
    bond = None  # Iterator for traversing list of bonds
    
    for bond in bonds:
        bond.exp = float(bond.exp)
        bond.yld = float(bond.yld)
        bond.ytd = float(bond.ytd)
        bond.mtd = float(bond.mtd)
        bond.qtd = float(bond.qtd)
        bond.t1 = float(bond.t1)
        bond.t3 = float(bond.t3)
        bond.t5 = float(bond.t5)

# Write bonds to spreadsheet
def write_bonds(sheet=None, bonds=[]):
    bond = None  # Iterator for traversing list of bonds
    values = {}  # Dictionary to map out locations of data on spreadsheet
    
    for bond in bonds:
        # Map data to locations on spreadsheet
        values = {6:(bond.exp / 100.0), 7:(bond.yld / 100.0), 15:bond.ytd, 16:bond.mtd, 17:bond.qtd, 18:bond.t1, 19:bond.t3, 20:bond.t5}

        # Write non-empty bond values to spreadsheet
        for column in list(values.keys()):
            sheet.cell(row=bond.row, column=column).value = values[column]
            sheet.cell(row=bond.row, column=column).font = Font(color=colors.BLACK, italic=False)

        # Write index bond values to spreadsheet in place of empty bond values
        for column in bond.index:
            sheet.cell(row=bond.row, column=column).font = Font(color=colors.BLUE, italic=True)

def main():
    file = None      # Text file storing path to Excel workbook
    path = ''        # String path to workbook
    workbook = None  # OpenPyXL workbook object
    sheet = None     # Active spreadsheet in workbook
    indexes = []     # List of index bonds
    bonds = []       # List of non-index bonds
    browser = None   # Selenium browser object
    bond = None      # Iterator for traversing list of non-index bonds
    index = None     # Iterator for traversing list of index bonds
    
    # Get path to workbook
    print('Getting spreadsheet path...')
    file = open('excelpath.txt')
    path = file.read()
    file.close()

    # Open workbook and store active sheet
    print('Opening workbook...')
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Get list of bonds and indexes from spreadsheet
    print('Getting list of bonds...')
    indexes = get_bonds(sheet, True)
    bonds = get_bonds(sheet, False)

    # Get bond values from Morningstar
    print('Getting index bond values...\n')
    browser = webdriver.Chrome()
    get_values(browser, indexes)
    print('Getting bond values...\n')
    get_values(browser, bonds)
    browser.quit()

    print('Filling in empty bond values with index data...')
    for bond in bonds:
        # Find corresponding index
        for index in indexes:
            if bond.row < index.row:
                break

        # Fill in empty bonds with index data and keep track of locations of empty bonds
        if bond.exp == '':
            bond.exp = index.exp
            bond.index.append(6)
        if bond.yld == '':
            bond.yld = index.yld
            bond.index.append(7)
        if bond.ytd == '':
            bond.ytd = index.ytd
            bond.index.append(15)
        if bond.mtd == '':
            bond.mtd = index.mtd
            bond.index.append(16)
        if bond.qtd == '':
            bond.qtd = index.qtd
            bond.index.append(17)
        if bond.t1 == '':
            bond.t1 = index.t1
            bond.index.append(18)
        if bond.t3 == '':
            bond.t3 = index.t3
            bond.index.append(19)
        if bond.t5 == '':
            bond.t5 = index.t5
            bond.index.append(20)

    # Convert index and bond values from strings to floats
    print('Converting bond values to decimal numbers...')
    to_floats(indexes)
    to_floats(bonds)

    # Modfiy path so original workbook isn't overwritten
    path = path.split('\\')
    del path[-1]
    path.append('output.xlsx')
    path = '\\'.join(path)
    
    # Save bonds to spreadsheet
    print('Saving bonds to spreadsheet...')
    write_bonds(sheet, indexes)
    write_bonds(sheet, bonds)
    workbook.save(path)

    print('Done!')

if __name__ == '__main__':
    main()
